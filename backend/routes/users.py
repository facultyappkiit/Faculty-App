from fastapi import APIRouter, HTTPException, status, Depends, UploadFile, File
from pydantic import BaseModel
from typing import List, Optional
import bcrypt
from datetime import datetime, time, timedelta
from openpyxl import load_workbook

from database import get_supabase
from models import UserResponse, UserUpdate, PushTokenUpdate, ClassScheduleItem
from middleware.auth import get_current_user, get_current_admin, get_super_admin, TokenData

router = APIRouter()


DAY_NAME_TO_INDEX = {
    "monday": 0,
    "mon": 0,
    "tuesday": 1,
    "tue": 1,
    "tues": 1,
    "wednesday": 2,
    "wed": 2,
    "thursday": 3,
    "thu": 3,
    "thur": 3,
    "thurs": 3,
    "friday": 4,
    "fri": 4,
    "saturday": 5,
    "sat": 5,
    "sunday": 6,
    "sun": 6,
}


def _normalize_header(value) -> str:
    if value is None:
        return ""
    return "".join(ch for ch in str(value).strip().lower() if ch.isalnum())


def _find_header_index(headers: list[str], options: list[str]) -> int | None:
    for index, header in enumerate(headers):
        for option in options:
            if option in header:
                return index
    return None


def _parse_day_of_week(value) -> int:
    if value is None:
        raise ValueError("Missing day value")

    if isinstance(value, (int, float)):
        numeric = int(value)
        if 0 <= numeric <= 6:
            return numeric
        if 1 <= numeric <= 7:
            return numeric - 1

    day_key = str(value).strip().lower()
    if day_key in DAY_NAME_TO_INDEX:
        return DAY_NAME_TO_INDEX[day_key]

    raise ValueError(f"Invalid day value: {value}")


def _parse_time_value(value) -> time:
    if value is None:
        raise ValueError("Missing time value")

    if isinstance(value, datetime):
        return value.time().replace(microsecond=0)
    if isinstance(value, time):
        return value.replace(microsecond=0)
    if isinstance(value, (int, float)):
        total_seconds = int(round(float(value) * 24 * 60 * 60)) % (24 * 60 * 60)
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60
        return time(hour=hours, minute=minutes, second=seconds)

    raw = str(value).strip()
    for fmt in ["%H:%M", "%H:%M:%S", "%I:%M %p", "%I:%M%p"]:
        try:
            return datetime.strptime(raw, fmt).time().replace(microsecond=0)
        except ValueError:
            continue

    raise ValueError(f"Invalid time value: {value}")


def _parse_duration_minutes(value) -> int:
    if value is None or str(value).strip() == "":
        raise ValueError("Missing duration value")

    if isinstance(value, (int, float)):
        duration = int(value)
    else:
        duration = int(float(str(value).strip()))

    if duration <= 0:
        raise ValueError(f"Invalid duration value: {value}")
    return duration


def _is_empty_schedule_cell(value) -> bool:
    if value is None:
        return True
    text = str(value).strip()
    if not text:
        return True
    # Treat placeholders like ---- and --- as empty.
    return all(ch == "-" for ch in text)


def _parse_time_slot_label(value) -> tuple[time, time] | None:
    raw = str(value or "").strip().replace(" ", "")
    if "-" not in raw:
        return None

    left, right = raw.split("-", 1)
    if not left or not right:
        return None

    if ":" in left or ":" in right:
        try:
            start_t = _parse_time_value(left)
            end_t = _parse_time_value(right)
            if end_t <= start_t:
                end_dt = datetime.combine(datetime.utcnow().date(), end_t) + timedelta(hours=12)
                end_t = end_dt.time().replace(microsecond=0)
            return start_t, end_t
        except ValueError:
            return None

    if not (left.isdigit() and right.isdigit()):
        return None

    start_hour = int(left)
    end_hour = int(right)

    # Timetable slots like 1-2, 2-3, ... are afternoon periods in this sheet style.
    if 1 <= start_hour <= 7:
        start_hour += 12
    if 1 <= end_hour <= 7:
        end_hour += 12
    if end_hour <= start_hour:
        end_hour += 12

    try:
        return time(start_hour % 24, 0, 0), time(end_hour % 24, 0, 0)
    except ValueError:
        return None


def _extract_matrix_schedule_rows(rows: list[tuple], file_name: str) -> list[dict]:
    if not rows:
        return []

    header_row_idx = None
    slot_columns: list[tuple[int, str, tuple[time, time]]] = []

    for idx, row in enumerate(rows):
        parsed_slots = []
        for col_idx, cell in enumerate(row):
            parsed = _parse_time_slot_label(cell)
            if parsed:
                parsed_slots.append((col_idx, str(cell).strip(), parsed))
        if len(parsed_slots) >= 3:
            header_row_idx = idx
            slot_columns = parsed_slots
            break

    if header_row_idx is None:
        return []

    records: list[dict] = []
    for row in rows[header_row_idx + 1:]:
        if not row:
            continue

        day_value = row[0] if len(row) > 0 else None
        try:
            day_of_week = _parse_day_of_week(day_value)
        except ValueError:
            continue

        for col_idx, _, (start_t, end_t) in slot_columns:
            if col_idx >= len(row):
                continue
            cell_value = row[col_idx]
            if _is_empty_schedule_cell(cell_value):
                continue

            cell_str = str(cell_value).strip()
            subject_text = cell_str
            classroom_text = None
            
            # Handle comma separated format: Section, Subject, Classroom, ...
            if ',' in cell_str:
                parts = [p.strip() for p in cell_str.split(',') if p.strip()]
                if len(parts) >= 3:
                    subject_text = f"{parts[0]} - {parts[1]}"
                    classroom_text = parts[2]
                elif len(parts) == 2:
                    subject_text = parts[0]
                    classroom_text = parts[1]
            elif '\n' in cell_str:
                lines = [line.strip() for line in cell_str.splitlines() if line.strip()]
                if lines:
                    subject_text = lines[0]
                if len(lines) > 1:
                    classroom_text = lines[1]

            if len(subject_text) > 120:
                subject_text = subject_text[:120]
            if classroom_text and len(classroom_text) > 50:
                classroom_text = classroom_text[:50]

            records.append(
                {
                    "day_of_week": day_of_week,
                    "start_time": start_t.strftime("%H:%M:%S"),
                    "end_time": end_t.strftime("%H:%M:%S"),
                    "subject": subject_text or None,
                    "classroom": classroom_text or None,
                    "source_file": file_name,
                }
            )

    return records


def _extract_columnar_schedule_rows(rows: list[tuple], file_name: str) -> list[dict]:
    records: list[dict] = []

    header_index = None
    day_idx = None
    start_idx = None
    end_idx = None
    duration_idx = None
    subject_idx = None
    classroom_idx = None

    for idx, row in enumerate(rows):
        normalized = [_normalize_header(cell) for cell in row]
        candidate_day_idx = _find_header_index(normalized, ["day", "weekday"])
        candidate_start_idx = _find_header_index(normalized, ["starttime", "start", "time"])
        candidate_end_idx = _find_header_index(normalized, ["endtime", "end"])
        candidate_duration_idx = _find_header_index(normalized, ["duration", "minutes", "mins"])
        candidate_subject_idx = _find_header_index(normalized, ["subject", "course", "class"])
        candidate_classroom_idx = _find_header_index(normalized, ["room", "classroom"])

        if candidate_day_idx is not None and candidate_start_idx is not None and (
            candidate_end_idx is not None or candidate_duration_idx is not None
        ):
            header_index = idx
            day_idx = candidate_day_idx
            start_idx = candidate_start_idx
            end_idx = candidate_end_idx
            duration_idx = candidate_duration_idx
            subject_idx = candidate_subject_idx
            classroom_idx = candidate_classroom_idx
            break

    if header_index is None or day_idx is None or start_idx is None:
        return []

    for row in rows[header_index + 1:]:
        if not any(cell is not None and str(cell).strip() != "" for cell in row):
            continue

        try:
            day_of_week = _parse_day_of_week(row[day_idx] if day_idx < len(row) else None)
            start_time = _parse_time_value(row[start_idx] if start_idx < len(row) else None)

            end_time = None
            if end_idx is not None and end_idx < len(row) and row[end_idx] is not None and str(row[end_idx]).strip() != "":
                end_time = _parse_time_value(row[end_idx])
            elif duration_idx is not None and duration_idx < len(row):
                duration = _parse_duration_minutes(row[duration_idx])
                end_dt = datetime.combine(datetime.utcnow().date(), start_time) + timedelta(minutes=duration)
                if end_dt.date() != datetime.utcnow().date():
                    raise ValueError("Duration crosses midnight")
                end_time = end_dt.time().replace(microsecond=0)

            if end_time is None or end_time <= start_time:
                raise ValueError("End time must be after start time")

            subject = None
            if subject_idx is not None and subject_idx < len(row) and row[subject_idx] is not None:
                subject = str(row[subject_idx]).strip() or None

            classroom = None
            if classroom_idx is not None and classroom_idx < len(row) and row[classroom_idx] is not None:
                classroom = str(row[classroom_idx]).strip() or None

            records.append(
                {
                    "day_of_week": day_of_week,
                    "start_time": start_time.strftime("%H:%M:%S"),
                    "end_time": end_time.strftime("%H:%M:%S"),
                    "subject": subject,
                    "classroom": classroom,
                    "source_file": file_name,
                }
            )
        except ValueError:
            continue

    return records


def _extract_schedule_rows(upload_file: UploadFile) -> list[dict]:
    file_name = upload_file.filename or "schedule.xlsx"
    lower_name = file_name.lower()

    if lower_name.endswith(".xls"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=".xls format is not supported yet. Please upload .xlsx"
        )

    if not lower_name.endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file type. Please upload an Excel .xlsx file"
        )

    try:
        upload_file.file.seek(0)
        workbook = load_workbook(upload_file.file, data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unable to read Excel file: {str(exc)}"
        )

    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            continue

        records = _extract_columnar_schedule_rows(rows, file_name)
        if not records:
            records = _extract_matrix_schedule_rows(rows, file_name)
        if records:
            return records

    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="No valid schedule rows found. Use either day/start/end columns or the timetable grid format"
    )


class AdminCreateUser(BaseModel):
    name: str
    email: str
    password: str
    department: Optional[str] = None
    phone: Optional[str] = None


def _is_valid_expo_push_token(token: str) -> bool:
    if not token:
        return False
    return token.startswith("ExponentPushToken[") or token.startswith("ExpoPushToken[")


@router.get("/", response_model=List[UserResponse])
async def get_all_users(current_admin: TokenData = Depends(get_current_admin)):
    """
    Get all registered faculty users.
    Admin only.
    """
    supabase = get_supabase()
    
    try:
        result = supabase.table("users")\
            .select("id, name, email, department, phone, email_verified, push_token, created_at")\
            .order("name")\
            .execute()
        
        return [
            UserResponse(
                id=user["id"],
                name=user["name"],
                email=user["email"],
                department=user.get("department"),
                phone=user.get("phone"),
                email_verified=user.get("email_verified", False),
                push_token=user.get("push_token"),
                created_at=user.get("created_at")
            )
            for user in result.data
        ]
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch users: {str(e)}"
        )


@router.post("/", response_model=UserResponse, status_code=status.HTTP_201_CREATED)
async def create_user_by_admin(user_data: AdminCreateUser, current_admin: TokenData = Depends(get_current_admin)):
    """
    Create a new user.
    Admin only.
    """
    supabase = get_supabase()
    
    # Validate email format
    if not user_data.email.endswith("@kiit.ac.in"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Email must be a KIIT email (@kiit.ac.in)"
        )
    
    try:
        # Check if email already exists
        existing = supabase.table("users").select("id").eq("email", user_data.email).execute()
        if existing.data:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="User with this email already exists"
            )
        
        # Hash the password
        hashed_password = bcrypt.hashpw(user_data.password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        
        # Create user
        result = supabase.table("users").insert({
            "name": user_data.name,
            "email": user_data.email,
            "password": hashed_password,
            "department": user_data.department,
            "phone": user_data.phone,
            "email_verified": True  # Admin-created users are pre-verified
        }).execute()
        
        if not result.data:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Failed to create user"
            )
        
        new_user = result.data[0]
        return UserResponse(
            id=new_user["id"],
            name=new_user["name"],
            email=new_user["email"],
            department=new_user.get("department"),
            phone=new_user.get("phone"),
            created_at=new_user.get("created_at")
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to create user: {str(e)}"
        )


@router.get("/{user_id}", response_model=UserResponse)
async def get_user(user_id: int, current_user: TokenData = Depends(get_current_user)):
    """
    Get a specific user by ID.
    Users can only view their own profile, admins can view any.
    """
    # Users can only view their own profile, admins can view any
    if current_user.token_type != "admin" and current_user.user_id != user_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not authorized to view this profile"
        )
    supabase = get_supabase()
    
    try:
        result = supabase.table("users")\
            .select("id, name, email, department, phone, created_at")\
            .eq("id", user_id)\
            .execute()
        
        if not result.data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="User not found"
            )
        
        user = result.data[0]
        return UserResponse(
            id=user["id"],
            name=user["name"],
            email=user["email"],
            department=user.get("department"),
            phone=user.get("phone"),
            created_at=user.get("created_at")
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch user: {str(e)}"
        )


@router.put("/{user_id}", response_model=UserResponse)
async def update_user(user_id: int, user_update: UserUpdate, current_user: TokenData = Depends(get_current_user)):
    """
    Update user profile information.
    Users can only update their own profile, admins can update any.
    """
    # Users can only update their own profile, admins can update any
    if current_user.token_type != "admin" and current_user.user_id != user_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not authorized to update this profile"
        )
    supabase = get_supabase()
    
    try:
        # Build update data (only non-None fields)
        update_data = {}
        if user_update.name is not None:
            update_data["name"] = user_update.name
        if user_update.department is not None:
            update_data["department"] = user_update.department
        if user_update.phone is not None:
            update_data["phone"] = user_update.phone
        
        if not update_data:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No fields to update"
            )
        
        result = supabase.table("users")\
            .update(update_data)\
            .eq("id", user_id)\
            .execute()
        
        if not result.data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="User not found"
            )
        
        user = result.data[0]
        return UserResponse(
            id=user["id"],
            name=user["name"],
            email=user["email"],
            department=user.get("department"),
            phone=user.get("phone"),
            created_at=user.get("created_at")
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to update user: {str(e)}"
        )


@router.put("/{user_id}/push-token")
async def update_push_token(user_id: int, token_update: PushTokenUpdate, current_user: TokenData = Depends(get_current_user)):
    """
    Update the push notification token for a user.
    Send JSON body: {"push_token": "ExponentPushToken[xxx]"}
    Users can only update their own push token.
    """
    # Users can only update their own push token
    if current_user.token_type != "admin" and current_user.user_id != user_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not authorized to update this push token"
        )
    supabase = get_supabase()
    
    token = token_update.push_token
    
    print(f"[PUSH-TOKEN] Received for user {user_id}: {token}")
    
    if not _is_valid_expo_push_token(token):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid Expo push token format"
        )
    
    try:
        result = supabase.table("users")\
            .update({"push_token": token})\
            .eq("id", user_id)\
            .execute()
        
        if not result.data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="User not found"
            )
        
        print(f"[PUSH-TOKEN] Saved for user {user_id}: {token[:40]}...")
        return {
            "message": "Push token updated successfully", 
            "user_id": user_id,
            "token_preview": token[:40] + "..."
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"[PUSH-TOKEN] Error: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to update push token: {str(e)}"
        )


@router.post("/{user_id}/push-token")
async def set_push_token_simple(user_id: int, push_token: str, current_user: TokenData = Depends(get_current_user)):
    """
    Simple endpoint to set push token via query parameter.
    Example: POST /api/users/1/push-token?push_token=ExponentPushToken[xxx]
    Users can only update their own push token.
    """
    # Users can only update their own push token
    if current_user.token_type != "admin" and current_user.user_id != user_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not authorized to update this push token"
        )
    supabase = get_supabase()
    
    print(f"[PUSH-TOKEN] POST received for user {user_id}: {push_token}")
    
    if not _is_valid_expo_push_token(push_token):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid Expo push token format"
        )
    
    try:
        result = supabase.table("users")\
            .update({"push_token": push_token})\
            .eq("id", user_id)\
            .execute()
        
        if not result.data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="User not found"
            )
        
        print(f"[PUSH-TOKEN] Saved for user {user_id}: {push_token[:40]}...")
        return {
            "message": "Push token updated successfully", 
            "user_id": user_id,
            "token_preview": push_token[:40] + "..."
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"[PUSH-TOKEN] Error: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to update push token: {str(e)}"
        )


@router.get("/{user_id}/push-token/status")
async def get_push_token_status(user_id: int, current_user: TokenData = Depends(get_current_user)):
    """
    Get push token registration status for a user.
    Useful for debugging notification registration from the app.
    Users can only check their own status, admins can check any.
    """
    # Users can only check their own status, admins can check any
    if current_user.token_type != "admin" and current_user.user_id != user_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not authorized to view this status"
        )
    supabase = get_supabase()

    try:
        result = supabase.table("users")\
            .select("id, name, push_token")\
            .eq("id", user_id)\
            .execute()

        if not result.data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="User not found"
            )

        user = result.data[0]
        token = user.get("push_token")
        return {
            "user_id": user["id"],
            "name": user.get("name"),
            "has_push_token": bool(token),
            "is_valid_expo_token": _is_valid_expo_push_token(token) if token else False,
            "token_preview": f"{token[:40]}..." if token else None,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to get push token status: {str(e)}"
        )


@router.post("/{user_id}/push-token/debug")
async def log_push_token_debug(user_id: int, payload: dict, current_user: TokenData = Depends(get_current_user)):
    """
    Receive client-side push registration debug state and print it in server logs.
    Users can only log debug for their own account.
    """
    # Users can only debug their own account
    if current_user.token_type != "admin" and current_user.user_id != user_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not authorized"
        )
    print(f"[PUSH-DEBUG] user_id={user_id} payload={payload}")
    return {"message": "Debug state logged"}


@router.post("/{user_id}/class-schedule/upload")
async def upload_class_schedule(
    user_id: int,
    schedule_file: Optional[UploadFile] = File(None),
    file: Optional[UploadFile] = File(None),
    current_user: TokenData = Depends(get_current_user),
):
    """
    Upload and replace a teacher's weekly class schedule from an Excel file.
    Users can upload only their own schedule, admins can upload for any user.
    """
    if current_user.token_type != "admin" and current_user.user_id != user_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not authorized to upload this schedule"
        )

    supabase = get_supabase()
    upload_file = schedule_file or file

    if upload_file is None:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Missing file. Send multipart field named 'schedule_file' or 'file'"
        )

    try:
        user_exists = supabase.table("users")\
            .select("id")\
            .eq("id", user_id)\
            .execute()

        if not user_exists.data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="User not found"
            )

        schedule_rows = _extract_schedule_rows(upload_file)
        payload = [
            {
                "teacher_id": user_id,
                "day_of_week": row["day_of_week"],
                "start_time": row["start_time"],
                "end_time": row["end_time"],
            }
            for row in schedule_rows
        ]

        supabase.table("teacher_class_schedules").delete().eq("teacher_id", user_id).execute()
        supabase.table("teacher_class_schedules").insert(payload).execute()

        days_covered = sorted({row["day_of_week"] for row in schedule_rows})
        return {
            "message": "Class schedule uploaded successfully",
            "user_id": user_id,
            "total_slots": len(schedule_rows),
            "days_covered": days_covered,
            "source_file": upload_file.filename,
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to upload class schedule: {str(e)}"
        )


@router.get("/{user_id}/class-schedule", response_model=List[ClassScheduleItem])
async def get_class_schedule(
    user_id: int,
    current_user: TokenData = Depends(get_current_user),
):
    """
    Get a teacher's weekly class schedule.
    Users can only view their own schedule, admins can view any.
    """
    if current_user.token_type != "admin" and current_user.user_id != user_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not authorized to view this schedule"
        )

    supabase = get_supabase()

    try:
        user_exists = supabase.table("users")\
            .select("id")\
            .eq("id", user_id)\
            .execute()

        if not user_exists.data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="User not found"
            )

        try:
            schedule_result = supabase.table("teacher_class_schedules")\
                .select("id, teacher_id, day_of_week, start_time, end_time, subject, classroom, substitute_request_id")\
                .eq("teacher_id", user_id)\
                .order("day_of_week", desc=False)\
                .order("start_time", desc=False)\
                .execute()
        except Exception as schedule_error:
            # Backward compatibility: older DB may not have substitute_request_id yet.
            if "substitute_request_id" not in str(schedule_error):
                raise

            schedule_result = supabase.table("teacher_class_schedules")\
                .select("id, teacher_id, day_of_week, start_time, end_time, subject, classroom")\
                .eq("teacher_id", user_id)\
                .order("day_of_week", desc=False)\
                .order("start_time", desc=False)\
                .execute()

        schedules = [
            ClassScheduleItem(
                id=item["id"],
                teacher_id=item["teacher_id"],
                day_of_week=item["day_of_week"],
                start_time=item["start_time"],
                end_time=item["end_time"],
                subject=item.get("subject"),
                classroom=item.get("classroom"),
                substitute_request_id=item.get("substitute_request_id"),
            )
            for item in (schedule_result.data or [])
        ]

        return schedules

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch class schedule: {str(e)}"
        )


@router.delete("/{user_id}")
async def delete_user(user_id: int, current_admin: TokenData = Depends(get_super_admin)):
    """
    Delete a user account.
    Super admin only.
    """
    supabase = get_supabase()
    
    try:
        check_result = supabase.table("users")\
            .select("id")\
            .eq("id", user_id)\
            .execute()
        
        if not check_result.data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="User not found"
            )
        
        supabase.table("users").delete().eq("id", user_id).execute()
        
        return {"message": "User deleted successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to delete user: {str(e)}"
        )
